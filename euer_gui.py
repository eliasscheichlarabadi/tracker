import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
from datetime import date, timedelta
import openpyxl
import csv
import json
import os
import re

# === KONFIGURATION ===
EXCEL_DATEI = "Umsatz 25.09 (2).xlsx"
SETTINGS_FILE = "settings.json"
UMSATZ_HISTORY_FILE = "umsatz_history.csv"
UMSAETZE_DIR = "umsaetze"

# === GRUNDSETUP ===
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("EÜR Rechner")
app.geometry("600x400")

# Funktion zum korrekten Beenden des Programms
def on_closing():
    app.quit()  # Beendet die Hauptschleife
    app.destroy()  # Zerstört das Fenster
    import sys
    sys.exit(0)  # Beendet das Programm vollständig

# Registriere die Funktion für das Schließen-Event
app.protocol("WM_DELETE_WINDOW", on_closing)

# === DATEN ===
heutiges_datum = date.today()
transaktionen = []
anfangsbestand = 0.0
firmenname = ""
firmen_liste = []
DB_CSV = "db.csv"
transaction_listbox = None
umsatz_history = []


def format_currency(value):
    """Format a value using German thousands separators and append €."""
    if value is None or value == "":
        return ""
    value = round(float(value), 2)
    if abs(value) >= 1000:
        formatted = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    else:
        formatted = f"{abs(value):.2f}".replace(".", ",")
    return f"{formatted} €"


def ensure_umsatz_dir():
    """Create the directory that stores saved revenue CSV files if missing."""
    try:
        os.makedirs(UMSAETZE_DIR, exist_ok=True)
    except Exception:
        pass


def sanitize_filename(name: str) -> str:
    """Replace characters that are invalid for filenames."""
    sanitized = re.sub(r"[\\/:*?\"<>|]", "_", name)
    sanitized = re.sub(r"_+", "_", sanitized)
    sanitized = sanitized.strip(" _")
    return sanitized or "umsatz"


def load_settings():
    """Load persisted settings such as the company list and last company."""
    global firmenname, firmen_liste
    if not os.path.exists(SETTINGS_FILE):
        return
    try:
        with open(SETTINGS_FILE, encoding="utf-8") as f:
            data = json.load(f)
        firmenname = data.get("zuletzt_verwendete_firma", "").strip()
        firmen_liste = sorted({name.strip() for name in data.get("firmen", []) if name.strip()})
    except Exception:
        firmenname = ""
        firmen_liste = []


def save_settings():
    """Persist settings like the company name list to disk."""
    data = {
        "zuletzt_verwendete_firma": firmenname,
        "firmen": sorted(firmen_liste),
    }
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_umsatz_history():
    """Load previously exported revenue summaries from disk."""
    global umsatz_history, firmen_liste
    if not os.path.exists(UMSATZ_HISTORY_FILE):
        return
    try:
        with open(UMSATZ_HISTORY_FILE, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")
            umsatz_history = []
            for row in reader:
                # Ensure expected keys exist for backwards compatibility
                entry = {
                    "erstellt_am": row.get("erstellt_am", ""),
                    "firma": row.get("firma", ""),
                    "bezugsdatum": row.get("bezugsdatum", ""),
                    "csv_datei": row.get("csv_datei") or row.get("dateiname", ""),
                    "einnahmen": row.get("einnahmen", ""),
                    "ausgaben": row.get("ausgaben", ""),
                    "gewinn": row.get("gewinn", ""),
                    "endbestand": row.get("endbestand", ""),
                }
                umsatz_history.append(entry)

            # Synchronisiere eventuell neue Firmennamen aus der Historie
            bekannte_firmen = set(firmen_liste)
            neue_firmen_gefunden = False
            for eintrag in umsatz_history:
                name = (eintrag.get("firma") or "").strip()
                if name and name not in bekannte_firmen:
                    firmen_liste.append(name)
                    bekannte_firmen.add(name)
                    neue_firmen_gefunden = True

            if neue_firmen_gefunden:
                firmen_liste.sort()
                save_settings()
    except Exception:
        umsatz_history = []


def save_umsatz_history():
    """Persist the list of exported revenue summaries to disk."""
    fieldnames = [
        "erstellt_am",
        "firma",
        "bezugsdatum",
        "csv_datei",
        "einnahmen",
        "ausgaben",
        "gewinn",
        "endbestand",
    ]
    try:
        with open(UMSATZ_HISTORY_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
            writer.writeheader()

            def sort_key(entry: dict) -> tuple:
                date_str = (entry.get("erstellt_am") or "")
                return (date_str, entry.get("csv_datei") or "")

            for entry in sorted(umsatz_history, key=sort_key, reverse=True):
                writer.writerow(entry)
    except Exception:
        pass


def record_umsatz_speicherung(csv_path, einnahmen, ausgaben, gewinn, endbestand):
    """Store metadata for the saved revenue report, replacing duplicates."""
    entry = {
        "erstellt_am": date.today().strftime("%Y-%m-%d"),
        "firma": firmenname,
        "bezugsdatum": aktuelles_datum.strftime("%Y-%m-%d"),
        "csv_datei": csv_path,
        "einnahmen": f"{einnahmen:.2f}",
        "ausgaben": f"{ausgaben:.2f}",
        "gewinn": f"{gewinn:.2f}",
        "endbestand": f"{endbestand:.2f}",
    }

    # Entferne bestehende Einträge mit derselben CSV-Datei, um Konflikte zu vermeiden
    umsatz_history[:] = [
        existing for existing in umsatz_history if existing.get("csv_datei") != csv_path
    ]
    umsatz_history.append(entry)

    def sort_key(item: dict) -> tuple:
        date_str = (item.get("erstellt_am") or "")
        return (date_str, item.get("csv_datei") or "")

    umsatz_history.sort(key=sort_key, reverse=True)
    save_umsatz_history()


def show_umsatz_history():
    """Display a window listing all exported revenue summaries."""
    if not umsatz_history:
        info_label.configure(text="ℹ️ Noch keine Umsätze gespeichert")
        return

    dialog = ctk.CTkToplevel(app)
    dialog.title("Gespeicherte Umsätze")
    dialog.geometry("720x360")

    option_frame = ctk.CTkFrame(dialog)
    option_frame.pack(fill="x", pady=(15, 5), padx=15)

    entries_with_labels = []
    for idx, entry in enumerate(umsatz_history, 1):
        label = (
            f"{idx}. {entry.get('firma', 'Unbekannt')} – {entry.get('bezugsdatum', '')}"
            f" (gespeichert am {entry.get('erstellt_am', '')})"
        )
        entries_with_labels.append((label, entry))

    selected_label = entries_with_labels[0][0]

    selection_var = ctk.StringVar(value=selected_label)
    entry_map = {label: data for label, data in entries_with_labels}

    def update_details(selected_value: str) -> None:
        entry = entry_map.get(selected_value)
        if not entry:
            return

        summary_lines = [
            f"Firma: {entry.get('firma', '')}",
            f"Bezugsdatum: {entry.get('bezugsdatum', '')}",
            f"Gespeichert am: {entry.get('erstellt_am', '')}",
            f"CSV-Datei: {entry.get('csv_datei', '')}",
            f"Einnahmen: {format_currency(entry.get('einnahmen'))}",
            f"Ausgaben: {format_currency(entry.get('ausgaben'))}",
            f"Gewinn: {format_currency(entry.get('gewinn'))}",
            f"Endbestand: {format_currency(entry.get('endbestand'))}",
            "",
            "Transaktionen:",
        ]

        details_box.configure(state="normal")
        details_box.delete("1.0", tk.END)
        details_box.insert(tk.END, "\n".join(summary_lines) + "\n")

        csv_path = entry.get("csv_datei", "")
        if csv_path and not os.path.isabs(csv_path):
            csv_path = os.path.join(UMSAETZE_DIR, csv_path)

        if csv_path and os.path.exists(csv_path):
            try:
                with open(csv_path, newline="", encoding="utf-8") as csv_file:
                    reader = csv.reader(csv_file, delimiter=";")
                    for row in reader:
                        if row:
                            details_box.insert(tk.END, " | ".join(row) + "\n")
                        else:
                            details_box.insert(tk.END, "\n")
            except Exception as exc:
                details_box.insert(tk.END, f"Fehler beim Lesen der CSV: {exc}\n")
        else:
            details_box.insert(tk.END, "CSV-Datei nicht gefunden.\n")

        details_box.configure(state="disabled")

    dropdown = ctk.CTkOptionMenu(
        option_frame,
        values=[label for label, _ in entries_with_labels],
        variable=selection_var,
        command=update_details,
        width=480,
    )
    dropdown.pack(side="left", padx=(0, 10), pady=10)

    close_button = ctk.CTkButton(option_frame, text="Schließen", command=dialog.destroy, width=120)
    close_button.pack(side="right", pady=10)

    details_box = ctk.CTkTextbox(dialog, height=220)
    details_box.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    update_details(selected_label)
    dialog.grab_set()


def create_new_umsatz():
    """Reset the current revenue data to start a new record."""
    if not messagebox.askyesno(
        "Neuen Umsatz anlegen",
        "Alle aktuellen Transaktionen löschen und neu beginnen?",
        parent=app,
    ):
        return

    global transaktionen, anfangsbestand, aktuelles_datum, firmenname
    transaktionen = []
    anfangsbestand = 0.0
    aktuelles_datum = date.today()
    firmenname = ""
    if os.path.exists(DB_CSV):
        try:
            os.remove(DB_CSV)
        except Exception:
            pass
    refresh_transaction_list()
    datum_anzeigen()
    ask_firma_if_needed(force=True)
    ask_anfangsbestand_if_needed(force=True)
    save_all_to_csv()
    info_label.configure(text="🆕 Neuer Umsatz vorbereitet. Bitte Transaktionen erfassen.")

# === FUNKTIONEN ===
def datum_anzeigen():
    datum_label.configure(text=aktuelles_datum.strftime("%d.%m.%Y"))

def datum_plus():
    global aktuelles_datum
    aktuelles_datum += timedelta(days=1)
    datum_anzeigen()

def datum_minus():
    global aktuelles_datum
    aktuelles_datum -= timedelta(days=1)
    datum_anzeigen()

def transaktion_hinzufügen():
    # accept comma as decimal separator
    betrag_text = betrag_entry.get().strip().replace(',', '.')
    try:
        betrag_raw = float(betrag_text) if betrag_text != "" else 0.0
    except ValueError:
        info_label.configure(text="❌ Ungültiger Betrag")
        return

    kategorie = kategorie_option.get()

    # Betrag als positiv/negativ speichern basierend auf Kategorie
    # Tagesumsatz Kasse ist Einnahme, alles andere Ausgabe
    if "Tagesumsatz Kasse" in kategorie:
        betrag = abs(betrag_raw)
    else:
        betrag = -abs(betrag_raw)

    trans = {
        "Datum": aktuelles_datum.strftime("%Y-%m-%d"),
        "Kategorie": kategorie,
        "Betrag": betrag
    }
    transaktionen.append(trans)

    # Komplette DB neu schreiben mit Anfangsbestand
    try:
        save_all_to_csv()  # Speichert Anfangsbestand + alle Transaktionen
    except Exception as e:
        info_label.configure(text=f"❌ Fehler beim Speichern in CSV: {e}")
        return

    # Formatiere Betrag mit Tausenderpunkten und Komma für die Anzeige
    betrag_abs = abs(betrag)
    if betrag_abs >= 1000:
        betrag_str = f"{betrag_abs:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    else:
        betrag_str = f"{betrag_abs:.2f}".replace(".", ",")
    info_label.configure(text=f"💾 Transaktion gespeichert ({kategorie}: {betrag_str} €)")
    betrag_entry.delete(0, "end")
    refresh_transaction_list()


def umsatz_speichern():
    """Persist the current Umsatz into a dedicated CSV file and log it."""
    if not firmenname:
        info_label.configure(text="❌ Bitte zuerst eine Firma auswählen")
        return

    ensure_umsatz_dir()

    einnahmen = sum(t["Betrag"] for t in transaktionen if t["Betrag"] > 0)
    ausgaben = sum(-t["Betrag"] for t in transaktionen if t["Betrag"] < 0)
    gewinn = einnahmen - ausgaben
    endbestand = anfangsbestand + gewinn

    base_name = sanitize_filename(f"{firmenname}_{aktuelles_datum.strftime('%Y-%m-%d')}")
    csv_filename = f"{base_name}.csv"
    csv_path = os.path.join(UMSAETZE_DIR, csv_filename)

    counter = 1
    while os.path.exists(csv_path):
        csv_filename = f"{base_name}_{counter}.csv"
        csv_path = os.path.join(UMSAETZE_DIR, csv_filename)
        counter += 1

    try:
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["Firma", firmenname])
            writer.writerow(["Bezugsdatum", aktuelles_datum.strftime("%Y-%m-%d")])
            writer.writerow(["Anfangsbestand", f"{anfangsbestand:.2f}"])
            writer.writerow([])
            writer.writerow(["Datum", "Kategorie", "Betrag"])
            for t in transaktionen:
                writer.writerow([t["Datum"], t["Kategorie"], f"{t['Betrag']:.2f}"])
            writer.writerow([])
            writer.writerow(["Summe Einnahmen", f"{einnahmen:.2f}"])
            writer.writerow(["Summe Ausgaben", f"{ausgaben:.2f}"])
            writer.writerow(["Gewinn", f"{gewinn:.2f}"])
            writer.writerow(["Endbestand", f"{endbestand:.2f}"])
    except Exception as exc:
        info_label.configure(text=f"❌ Fehler beim Speichern des Umsatzes: {exc}")
        return

    record_umsatz_speicherung(csv_filename, einnahmen, ausgaben, gewinn, endbestand)
    info_label.configure(text=f"✅ Umsatz gespeichert: {csv_filename}")

def exportieren():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EÜR"

        # Styles definieren
        header_font = openpyxl.styles.Font(name='Arial', bold=True, size=10)
        normal_font = openpyxl.styles.Font(name='Arial', size=10)
        money_format = '@'  # Text-Format für vorformatierte Zahlen
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        header_fill = openpyxl.styles.PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        grey_fill = openpyxl.styles.PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Firmenname in der ersten Zeile
        ws.append([firmenname])
        ws.append([])  # Leerzeile
        
        # Anfangsbestand in der dritten Zeile
        ws.append(["", "", "", "Anfangsbestand:", format_currency(anfangsbestand)])
        anfang_row = ws.max_row
        for col in range(4, 6):
            cell = ws.cell(row=anfang_row, column=col)
            cell.font = header_font
            cell.border = thin_border
            if col == 4:  # "Anfangsbestand:" Text
                cell.fill = grey_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            elif col == 5:  # Betrag
                cell.number_format = money_format
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # Kopfzeile
        header_row = ws.max_row + 1
        ws.append(["Beleg-Nr.", "Datum", "Transaktion", "Einnahmen", "Ausgaben"])
        for col in range(1, 6):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            if col in [4, 5]:  # Geldbeträge rechtsbündig
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            else:  # Rest zentriert
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Transaktionen einfügen
        for idx, t in enumerate(transaktionen, 1):
            betrag = t["Betrag"]
            # Formatiere die Beträge
            einnahme = format_currency(abs(betrag)) if betrag > 0 else ""
            ausgabe = format_currency(abs(betrag)) if betrag < 0 else ""
            row = ws.append([
                idx,  # Beleg-Nr.
                t["Datum"].replace("-", "/"),  # Datum im Format DD/MM/YYYY
                t["Kategorie"].split(" ", 1)[1],  # Kategorie ohne Emoji
                einnahme,
                ausgabe
            ])
            
            # Formatierung der Zeile
            current_row = ws.max_row
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = normal_font
                cell.border = thin_border  # Vollständiger Rand für alle Zellen
                if col in [4, 5]:  # Geldbeträge
                    cell.number_format = money_format
                    cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                elif col == 1:  # Beleg-Nr.
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                elif col == 2:  # Datum
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                else:  # Transaktion
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            
            # Beleg-Nr. zentrieren
            ws.cell(row=current_row, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

        # Summen
        einnahmen = sum(t["Betrag"] for t in transaktionen if t["Betrag"] > 0)
        ausgaben = sum(-t["Betrag"] for t in transaktionen if t["Betrag"] < 0)
        gewinn = einnahmen - ausgaben
        endbestand = anfangsbestand + gewinn

        ws.append([])
        # Untere Kante für letzte Transaktionszeile
        last_data_row = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=last_data_row, column=col)
            current_border = cell.border
            cell.border = openpyxl.styles.Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=openpyxl.styles.Side(style='thin')
            )

        ws.append([])  # Leerzeile vor Summen
        
        # Summen einfügen mit angepasster Formatierung
        # Gesamtzeile mit Einnahmen und Ausgaben nebeneinander
        ws.append(["", "", "Gesamt:", format_currency(einnahmen), format_currency(ausgaben)])
        current_row = ws.max_row
        
        # Gesamttext mit grauem Hintergrund
        gesamt_cell = ws.cell(row=current_row, column=3)
        gesamt_cell.font = header_font
        gesamt_cell.border = thin_border
        gesamt_cell.fill = grey_fill
        gesamt_cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        
        # Einnahmen- und Ausgabenbetrag ohne Fülleffekt
        for col in [4, 5]:  # Spalten für Einnahmen und Ausgaben
            cell = ws.cell(row=current_row, column=col)
            cell.font = header_font
            cell.border = thin_border
            cell.number_format = money_format
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # Endbestand (Text mit grauem Fülleffekt, Betrag ohne)
        ws.append(["", "", "", "Endbestand:", format_currency(endbestand)])
        current_row = ws.max_row
        for col in range(4, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = header_font
            cell.border = thin_border
            if col == 4:  # "Endbestand:" Text
                cell.fill = grey_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            elif col == 5:  # Geldbetrag
                cell.number_format = money_format
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # Optimierte Spaltenbreiten
        column_widths = {
            1: 12,  # Beleg-Nr.
            2: 12,  # Datum
            3: 40,  # Transaktion (breiter für lange Kategorienamen)
            4: 18,  # Einnahmen
            5: 18   # Ausgaben
        }
        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Stelle sicher, dass jede Zelle Arial in Schriftgröße 10 verwendet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.font = cell.font.copy(name="Arial", size=10)

        # Build filename (e.g. 'Umsatz 25.11.xlsx')
        filename = f"Umsatz {aktuelles_datum.strftime('%y.%m')}.xlsx"
        wb.save(filename)

        info_label.configure(text=f"📤 Export erfolgreich: {filename}")
    except Exception as e:
        info_label.configure(text=f"❌ Fehler beim Export: {e}")


def load_db():
    """Load transactions and anfangsbestand from DB_CSV if present."""
    global anfangsbestand, transaktionen, firmenname
    if not os.path.exists(DB_CSV):
        return
    try:
        with open(DB_CSV, newline="", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter=";")
            rows = list(reader)
        if not rows:
            return
        transaktionen = []
        anfangsbestand = 0.0
        firmenname_local = ""
        for row in rows:
            if not row:
                continue
            key = row[0].strip().lower()
            if key == "firma" and len(row) > 1:
                firmenname_local = row[1].strip()
            elif key == "anfangsbestand" and len(row) > 1:
                try:
                    anfangsbestand = float(row[1].replace(',', '.'))
                except Exception:
                    anfangsbestand = 0.0
            elif len(row) >= 3:
                try:
                    tdate = row[0]
                    k = row[1]
                    b = float(row[2].replace(',', '.'))
                    transaktionen.append({"Datum": tdate, "Kategorie": k, "Betrag": b})
                except Exception:
                    continue

        if firmenname_local:
            firmenname = firmenname_local
            if firmenname not in firmen_liste and firmenname:
                firmen_liste.append(firmenname)
                firmen_liste.sort()
                save_settings()
    except Exception:
        return


def save_all_to_csv():
    """Write the entire DB (Anfangsbestand header if present + all transactions)."""
    try:
        with open(DB_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=';')
            if firmenname:
                writer.writerow(["Firma", firmenname])
            # write anfangsbestand as first line
            writer.writerow(["Anfangsbestand", f"{anfangsbestand:.2f}"])
            for t in transaktionen:
                writer.writerow([t["Datum"], t["Kategorie"], f"{t['Betrag']:.2f}"])
    except Exception:
        return


def refresh_transaction_list():
    """Refresh the Listbox content from transaktionen."""
    transaction_listbox.delete(0, tk.END)
    for idx, t in enumerate(transaktionen):
        # Formatiere Betrag mit Tausenderpunkten und Komma
        betrag = abs(t['Betrag'])
        if betrag >= 1000:
            betrag_str = f"{betrag:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        else:
            betrag_str = f"{betrag:.2f}".replace(".", ",")
        
        # Feste Breiten für die Spalten
        datum = t['Datum'].ljust(10)  # Datum hat immer 10 Zeichen
        kategorie = t['Kategorie'].ljust(30)  # Kategorie auf 30 Zeichen auffüllen
        betrag_str = f"{betrag_str} €".rjust(15)  # Betrag rechtsbündig, 15 Zeichen
        
        display = f"{datum} | {kategorie} | {betrag_str}"
        transaction_listbox.insert(tk.END, display)
    
    # Scrolle zur letzten Transaktion
    transaction_listbox.see(tk.END)
    transaction_listbox.selection_clear(0, tk.END)  # Entferne eventuell vorhandene Auswahl


def delete_selected_transaction():
    sel = transaction_listbox.curselection()
    if not sel:
        info_label.configure(text="❌ Keine Transaktion ausgewählt")
        return
    idx = sel[0]
    # remove from memory
    try:
        removed = transaktionen.pop(idx)
    except Exception:
        info_label.configure(text="❌ Fehler beim Löschen")
        return
    # rewrite CSV preserving anfangsbestand
    save_all_to_csv()
    refresh_transaction_list()
    # Formatiere Betrag für die Anzeige
    betrag_abs = abs(removed['Betrag'])
    if betrag_abs >= 1000:
        betrag_str = f"{betrag_abs:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    else:
        betrag_str = f"{betrag_abs:.2f}".replace(".", ",")
    info_label.configure(text=f"🗑️ Transaktion gelöscht: {removed['Kategorie']} {betrag_str} €")


def ask_firma_if_needed(force: bool = False):
    """Fragt nach dem Firmennamen, sofern nötig oder erzwungen."""
    global firmenname
    if firmenname and not force:
        return

    dialog = ctk.CTkToplevel(app)
    dialog.title("Firma auswählen")
    dialog.geometry("420x220")

    label = ctk.CTkLabel(dialog, text="Für welche Firma soll der Umsatz erfasst werden?", wraplength=360)
    label.pack(pady=(20, 10))

    options = sorted(firmen_liste)
    new_option = "➕ Neue Firma anlegen"
    options.append(new_option)

    if force and new_option in options:
        default_value = new_option
    elif firmenname and firmenname in options:
        default_value = firmenname
    else:
        default_value = options[0]

    selection_var = ctk.StringVar(value=default_value)

    entry = ctk.CTkEntry(dialog, width=260)
    if default_value == new_option:
        entry.configure(state="normal")
    else:
        entry.insert(0, default_value)
        entry.configure(state="disabled")

    def on_select(value: str) -> None:
        if value == new_option:
            entry.configure(state="normal")
            entry.delete(0, tk.END)
            entry.focus_set()
        else:
            entry.configure(state="normal")
            entry.delete(0, tk.END)
            entry.insert(0, value)
            entry.configure(state="disabled")

    dropdown = ctk.CTkOptionMenu(dialog, values=options, variable=selection_var, command=on_select, width=260)
    dropdown.pack(pady=(5, 10))

    entry.pack(pady=(0, 10))

    def submit() -> None:
        global firmenname
        choice = selection_var.get()
        if choice == new_option:
            firma = entry.get().strip()
        else:
            firma = choice.strip()

        if not firma:
            info_label.configure(text="❌ Bitte Firmennamen eingeben")
            return

        firmenname = firma
        if firmenname not in firmen_liste:
            firmen_liste.append(firmenname)
            firmen_liste.sort()
        save_settings()
        save_all_to_csv()
        info_label.configure(text=f"Firma gesetzt: {firmenname}")
        dialog.destroy()

    submit_btn = ctk.CTkButton(dialog, text="OK", command=submit, width=140)
    submit_btn.pack(pady=(5, 15))

    on_select(selection_var.get())

    dialog.grab_set()
    app.wait_window(dialog)


def ask_anfangsbestand_if_needed(force: bool = False):
    """Ask user for Anfangsbestand if not already set from CSV or forced."""
    global anfangsbestand
    if anfangsbestand != 0.0 and not force:
        return

    dialog = ctk.CTkToplevel(app)
    dialog.title("Anfangsbestand")
    dialog.geometry("320x140")
    label = ctk.CTkLabel(dialog, text="Anfangsbestand (€):")
    label.pack(pady=(20, 8))
    entry = ctk.CTkEntry(dialog)
    entry.pack(pady=5)

    def submit() -> None:
        nonlocal_entry = entry.get()
        try:
            val = float(nonlocal_entry) if nonlocal_entry.strip() != "" else 0.0
        except ValueError:
            info_label.configure(text="❌ Ungültiger Anfangsbestand")
            return

        global anfangsbestand
        anfangsbestand = val

        # Bewahre bestehende Transaktionen ohne doppelte Kopfzeilen
        remaining_rows = []
        if os.path.exists(DB_CSV):
            try:
                with open(DB_CSV, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f, delimiter=";")
                    for row in reader:
                        if not row:
                            continue
                        key = row[0].strip().lower()
                        if key in {"anfangsbestand", "firma"}:
                            continue
                        remaining_rows.append(row)
            except Exception:
                remaining_rows = []

        try:
            with open(DB_CSV, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=';')
                if firmenname:
                    writer.writerow(["Firma", firmenname])
                writer.writerow(["Anfangsbestand", f"{anfangsbestand:.2f}"])
                writer.writerows(remaining_rows)
        except Exception as exc:
            info_label.configure(text=f"❌ Fehler beim Schreiben der DB: {exc}")
            dialog.destroy()
            return

        if anfangsbestand >= 1000:
            bestand_str = f"{anfangsbestand:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        else:
            bestand_str = f"{anfangsbestand:.2f}".replace(".", ",")
        info_label.configure(text=f"Anfangsbestand gesetzt: {bestand_str} €")
        dialog.destroy()

    submit_btn = ctk.CTkButton(dialog, text="OK", command=submit)
    submit_btn.pack(pady=12)
    dialog.grab_set()
    app.wait_window(dialog)

# === GUI ELEMENTE ===
aktuelles_datum = heutiges_datum

datum_frame = ctk.CTkFrame(app)
datum_frame.pack(pady=10)

minus_button = ctk.CTkButton(datum_frame, text="◀", width=40, command=datum_minus)
minus_button.pack(side="left", padx=5)

datum_label = ctk.CTkLabel(datum_frame, text="")
datum_label.pack(side="left", padx=10)
datum_anzeigen()

plus_button = ctk.CTkButton(datum_frame, text="▶", width=40, command=datum_plus)
plus_button.pack(side="left", padx=5)

# Kategorie Auswahl mit einheitlicher Einrückung
kategorien = [
    "💰  Tagesumsatz Kasse",
    "⛽  Tankbeleg",
    "🧹  Rechnung Teppichreinigung",
    "💶  Bargeldeinzahlung",
    "👤  Bargeldeinzahlung - Privat",
    "📊  Buchhaltungsservice",
    "🛍️  Wareneinkauf"
]
kategorie_option = ctk.CTkOptionMenu(app, values=kategorien, width=250)
kategorie_option.pack(pady=10)

# Betrag Eingabe
betrag_entry = ctk.CTkEntry(app, placeholder_text="Betrag (€)")
betrag_entry.pack(pady=10)

# Enter-Taste im Betrags-Feld führt Hinzufügen aus
def on_enter_pressed(event):
    transaktion_hinzufügen()
betrag_entry.bind("<Return>", on_enter_pressed)

# Buttons
add_button = ctk.CTkButton(app, text="Transaktion hinzufügen", command=transaktion_hinzufügen)
add_button.pack(pady=5)

export_button = ctk.CTkButton(app, text="📤 In Excel exportieren", fg_color="green", command=exportieren)
export_button.pack(pady=10)

save_umsatz_button = ctk.CTkButton(app, text="💾 Umsatz speichern", fg_color="#16a085", command=umsatz_speichern)
save_umsatz_button.pack(pady=5)

history_button = ctk.CTkButton(app, text="📂 Umsätze anzeigen", command=show_umsatz_history)
history_button.pack(pady=5)

new_umsatz_button = ctk.CTkButton(app, text="🆕 Neuen Umsatz anlegen", fg_color="#2980b9", command=create_new_umsatz)
new_umsatz_button.pack(pady=5)

# Info Label
info_label = ctk.CTkLabel(app, text="")
info_label.pack(pady=10)

# Transaction list and delete button
transactions_frame = ctk.CTkFrame(app)
transactions_frame.pack(pady=(5, 10), fill="both", expand=False)

# Use classic Tk Listbox inside the CTk frame for simplicity
listbox_container = tk.Frame(transactions_frame)
listbox_container.pack(side="left", fill="both", expand=True, padx=(10, 0))

transaction_listbox = tk.Listbox(listbox_container, height=8, width=60)
transaction_listbox.pack(side="left", fill="both", expand=True)

scrollbar = tk.Scrollbar(listbox_container, orient="vertical", command=transaction_listbox.yview)
scrollbar.pack(side="right", fill="y")
transaction_listbox.config(yscrollcommand=scrollbar.set)

delete_button = ctk.CTkButton(transactions_frame, text="Transaktion löschen", fg_color="#e74c3c", command=delete_selected_transaction)
delete_button.pack(side="left", padx=10)

# Load existing DB and ask for firma and Anfangsbestand if needed before starting
load_settings()
ensure_umsatz_dir()
load_umsatz_history()
load_db()
refresh_transaction_list()
ask_firma_if_needed()
ask_anfangsbestand_if_needed()
# Formatiere Anfangsbestand für die Anzeige
if anfangsbestand >= 1000:
    bestand_str = f"{anfangsbestand:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
else:
    bestand_str = f"{anfangsbestand:.2f}".replace(".", ",")
firmeninfo = f" | Firma: {firmenname}" if firmenname else ""
info_label.configure(text=f"Anfangsbestand: {bestand_str} € | geladene Transaktionen: {len(transaktionen)}{firmeninfo}")

app.mainloop()
